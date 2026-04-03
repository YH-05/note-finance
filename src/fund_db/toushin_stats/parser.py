"""Parser for Investment Trust Association statistics Excel files.

Reads XLSX files using openpyxl and converts rows into Pydantic
model instances. Supports four report types: B-1 (asset flow),
B-2 (product class), B-3 (management company), and A-2 (overall status).

Classes
-------
ToushinStatsParser
    Parses statistics Excel files into model instances.

Examples
--------
>>> from pathlib import Path
>>> parser = ToushinStatsParser()
>>> # records = parser.parse_b1(Path(".tmp/toushin_B1_shisan_zougen.xlsx"))
"""

from __future__ import annotations

import contextlib
import math
import re
from datetime import datetime
from typing import TYPE_CHECKING, Any

import openpyxl

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

from fund_db._logging import get_logger
from fund_db.exceptions import ParseError
from fund_db.toushin_stats.models import (
    AssetFlowRecord,
    ManagementCompanyRecord,
    OverallStatusRecord,
    ProductClassRecord,
)

logger = get_logger(__name__, module="toushin_stats_parser")

_EMPTY_MARKERS = {"", "-", "None", "N/A"}
_YM_SLASH_RE = re.compile(r"^(\d{4})/(\d{1,2})$")
_YM_JP_RE = re.compile(r"^(\d{4})年(\d{1,2})月$")


def _to_float(value: Any) -> float | None:
    """Convert a cell value to float or None.

    Parameters
    ----------
    value : Any
        Raw cell value from openpyxl.

    Returns
    -------
    float | None
        Numeric value as float, or None if empty/invalid.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        is_invalid = isinstance(value, float) and (
            math.isnan(value) or math.isinf(value)
        )
        return None if is_invalid else float(value)
    if not isinstance(value, str):
        return None
    stripped = value.strip()
    if stripped in _EMPTY_MARKERS:
        return None
    try:
        return float(stripped.replace(",", ""))
    except ValueError:
        return None


def _parse_ym_parts(year: int, month: int) -> str | None:
    """Validate and format year/month parts into "YYYY-MM"."""
    if 1 <= month <= 12 and 1900 <= year <= 2100:
        return f"{year:04d}-{month:02d}"
    return None


def _to_year_month(value: Any) -> str | None:
    """Convert a cell value to "YYYY-MM" format string or None.

    Handles various formats: "2024/1", "2024-01", "2024年1月",
    datetime objects, etc.

    Parameters
    ----------
    value : Any
        Raw cell value from openpyxl.

    Returns
    -------
    str | None
        Year-month string in "YYYY-MM" format, or None if invalid.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")

    text = str(value).strip()
    if not text or text in _EMPTY_MARKERS:
        return None

    # Try "YYYY-MM" format
    result: str | None = None
    if len(text) == 7 and text[4] == "-":
        with contextlib.suppress(ValueError):
            result = _parse_ym_parts(int(text[:4]), int(text[5:7]))

    # Try "YYYY/M" or "YYYY/MM" format
    if result is None:
        m_slash = _YM_SLASH_RE.match(text)
        if m_slash:
            result = _parse_ym_parts(int(m_slash.group(1)), int(m_slash.group(2)))

    # Try "YYYY年M月" format
    if result is None:
        m_jp = _YM_JP_RE.match(text)
        if m_jp:
            result = _parse_ym_parts(int(m_jp.group(1)), int(m_jp.group(2)))

    return result


class ToushinStatsParser:
    """Parses Investment Trust Association statistics Excel files.

    Implements parsing for four report types:

    - B-1: Asset flow (monthly time series)
    - B-2: Product class breakdown (multi-sheet)
    - B-3: Management company breakdown
    - A-2: Overall market status

    Examples
    --------
    >>> parser = ToushinStatsParser()
    """

    def _find_header_row_b1(
        self,
        sheet: Worksheet,
    ) -> tuple[int, dict[int, str]]:
        """Find the header row for B-1 asset flow data.

        Searches for rows containing known column headers such as
        "pure asset total amount" or "subscription" or "redemption" keywords.

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet to search.

        Returns
        -------
        tuple[int, dict[int, str]]
            Header row number (1-based) and mapping from
            column index (0-based) to field name.

        Raises
        ------
        ParseError
            If no header row is found.
        """
        # Known header keywords for B-1
        header_keywords: dict[str, str] = {
            "純資産": "net_assets",
            "設定": "inflow",
            "解約": "outflow",
            "純増減": "net_flow",
            "増減": "net_flow",
        }

        for row_idx, row in enumerate(
            sheet.iter_rows(max_row=30, values_only=False), start=1
        ):
            matched_fields = self._match_header_cells(row, header_keywords)
            if len(matched_fields) >= 2:
                logger.debug(
                    "B-1 header row found",
                    row=row_idx,
                    matched_columns=len(matched_fields),
                )
                return row_idx, matched_fields

        msg = "B-1 header row not found in worksheet"
        raise ParseError(
            msg,
            source=sheet.title or "unknown",
            reason="No matching header row for B-1 data",
        )

    @staticmethod
    def _match_header_cells(
        row: tuple[Any, ...],
        keywords: dict[str, str],
    ) -> dict[int, str]:
        """Match cells in a row against header keyword patterns.

        Parameters
        ----------
        row : tuple
            Row of cells from openpyxl.
        keywords : dict[str, str]
            Mapping from keyword to field name.

        Returns
        -------
        dict[int, str]
            Column index (0-based) -> field name for matched cells.
        """
        matched: dict[int, str] = {}
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip()
            for keyword, field_name in keywords.items():
                if keyword in text and field_name not in matched.values():
                    matched[cell.column - 1] = field_name
                    break
        return matched

    def _detect_year_month_column(
        self,
        sheet: Worksheet,
        header_row: int,
        col_map: dict[int, str],
    ) -> int | None:
        """Detect which column contains year-month values.

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet.
        header_row : int
            1-based row number of the header.
        col_map : dict[int, str]
            Known column mappings (to exclude).

        Returns
        -------
        int | None
            0-based column index, or None if not found.
        """
        for row in sheet.iter_rows(
            min_row=header_row + 1,
            max_row=header_row + 5,
            values_only=False,
        ):
            for cell in row:
                if cell.column is None:
                    continue
                col_idx = cell.column - 1
                if col_idx not in col_map:
                    ym = _to_year_month(cell.value)
                    if ym is not None:
                        return col_idx
        return None

    def _extract_year_month(
        self,
        row: tuple[Any, ...],
        year_month_col: int | None,
    ) -> str | None:
        """Extract year-month from a data row.

        Parameters
        ----------
        row : tuple
            Row of cells.
        year_month_col : int | None
            Detected year-month column index.

        Returns
        -------
        str | None
            Year-month in "YYYY-MM" format, or None.
        """
        if year_month_col is not None and year_month_col < len(row):
            ym_value = _to_year_month(row[year_month_col].value)
            if ym_value is not None:
                return ym_value
        # Fallback to first column
        if len(row) > 0:
            return _to_year_month(row[0].value)
        return None

    def parse_b1(self, path: Path) -> list[AssetFlowRecord]:
        """Parse B-1 (asset flow) Excel file.

        Reads the first sheet of the workbook, finds the header row,
        and converts each data row to an ``AssetFlowRecord``.

        Parameters
        ----------
        path : Path
            Path to the B-1 XLSX file.

        Returns
        -------
        list[AssetFlowRecord]
            Parsed asset flow records.

        Raises
        ------
        ParseError
            If the file cannot be read or parsed.
        """
        logger.info("Parsing B-1 asset flow Excel", path=str(path))
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise ParseError(
                f"Failed to open workbook: {exc}",
                source=str(path),
                reason=str(exc),
            ) from exc

        try:
            sheet = wb.worksheets[0] if wb.worksheets else None
            if sheet is None:
                raise ParseError(
                    "No active sheet found",
                    source=str(path),
                    reason="Workbook has no active sheet",
                )

            header_row, col_map = self._find_header_row_b1(sheet)
            year_month_col = self._detect_year_month_column(sheet, header_row, col_map)
            records = self._read_b1_rows(sheet, header_row, col_map, year_month_col)
        finally:
            wb.close()

        logger.info(
            "B-1 parsing complete",
            record_count=len(records),
            path=str(path),
        )
        return records

    def _read_b1_rows(
        self,
        sheet: Worksheet,
        header_row: int,
        col_map: dict[int, str],
        year_month_col: int | None,
    ) -> list[AssetFlowRecord]:
        """Read data rows from B-1 sheet and convert to records.

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet.
        header_row : int
            1-based header row number.
        col_map : dict[int, str]
            Column index -> field name mapping.
        year_month_col : int | None
            Detected year-month column index.

        Returns
        -------
        list[AssetFlowRecord]
            Parsed records.
        """
        records: list[AssetFlowRecord] = []
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=False),
            start=header_row + 1,
        ):
            ym_value = self._extract_year_month(row, year_month_col)
            if ym_value is None:
                continue

            row_data: dict[str, float | None] = {}
            for col_idx, field_name in col_map.items():
                if col_idx < len(row):
                    row_data[field_name] = _to_float(row[col_idx].value)

            if all(v is None for v in row_data.values()):
                continue

            try:
                record = AssetFlowRecord(
                    year_month=ym_value,
                    net_assets=row_data.get("net_assets"),
                    inflow=row_data.get("inflow"),
                    outflow=row_data.get("outflow"),
                    net_flow=row_data.get("net_flow"),
                )
                records.append(record)
            except Exception as exc:
                logger.warning(
                    "Failed to parse B-1 row",
                    row=row_idx,
                    error=str(exc),
                )
        return records

    # ------------------------------------------------------------------
    # B-2 (product class) parsing
    # ------------------------------------------------------------------

    def _find_header_row_b2(
        self,
        sheet: Worksheet,
    ) -> tuple[int, dict[int, str]] | None:
        """Find the header row for B-2 product class data.

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet to search.

        Returns
        -------
        tuple[int, dict[int, str]] | None
            Header row number (1-based) and column mapping,
            or None if no header row found.
        """
        header_keywords: dict[str, str] = {
            "純資産": "net_assets",
            "ファンド": "fund_count",
            "本数": "fund_count",
        }

        # First pass: look for rows matching at least 2 keywords
        for row_idx, row in enumerate(
            sheet.iter_rows(max_row=30, values_only=False), start=1
        ):
            matched_fields = self._match_header_cells(row, header_keywords)
            if len(matched_fields) >= 2:
                logger.debug(
                    "B-2 header row found (multi-match)",
                    sheet=sheet.title,
                    row=row_idx,
                    matched_columns=len(matched_fields),
                )
                return row_idx, matched_fields

        # Second pass: accept a single match if it comes from a row
        # containing multiple cells (not a title row)
        for row_idx, row in enumerate(
            sheet.iter_rows(max_row=30, values_only=False), start=1
        ):
            matched_fields = self._match_header_cells(row, header_keywords)
            non_empty_cells = sum(1 for c in row if c.value is not None)
            if len(matched_fields) >= 1 and non_empty_cells >= 2:
                logger.debug(
                    "B-2 header row found (single-match with context)",
                    sheet=sheet.title,
                    row=row_idx,
                    matched_columns=len(matched_fields),
                )
                return row_idx, matched_fields

        return None

    def parse_b2(self, path: Path) -> list[ProductClassRecord]:
        """Parse B-2 (product class) Excel file.

        Reads all sheets from the workbook, using each sheet name
        as the ``product_class`` field. Each sheet should contain
        year-month rows with net_assets and fund_count columns.

        Parameters
        ----------
        path : Path
            Path to the B-2 XLSX file.

        Returns
        -------
        list[ProductClassRecord]
            Parsed product class records from all sheets.

        Raises
        ------
        ParseError
            If the file cannot be read.
        """
        logger.info("Parsing B-2 product class Excel", path=str(path))
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise ParseError(
                f"Failed to open workbook: {exc}",
                source=str(path),
                reason=str(exc),
            ) from exc

        records: list[ProductClassRecord] = []
        try:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                header_result = self._find_header_row_b2(sheet)
                if header_result is None:
                    logger.debug(
                        "Skipping sheet without recognizable header",
                        sheet=sheet_name,
                    )
                    continue

                header_row, col_map = header_result
                year_month_col = self._detect_year_month_column(
                    sheet, header_row, col_map
                )
                sheet_records = self._read_b2_rows(
                    sheet, sheet_name, header_row, col_map, year_month_col
                )
                records.extend(sheet_records)
        finally:
            wb.close()

        logger.info(
            "B-2 parsing complete",
            record_count=len(records),
            path=str(path),
        )
        return records

    def _read_b2_rows(
        self,
        sheet: Worksheet,
        sheet_name: str,
        header_row: int,
        col_map: dict[int, str],
        year_month_col: int | None,
    ) -> list[ProductClassRecord]:
        """Read data rows from a B-2 sheet and convert to records.

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet.
        sheet_name : str
            Sheet name to use as product_class.
        header_row : int
            1-based header row number.
        col_map : dict[int, str]
            Column index -> field name mapping.
        year_month_col : int | None
            Detected year-month column index.

        Returns
        -------
        list[ProductClassRecord]
            Parsed records for this sheet.
        """
        records: list[ProductClassRecord] = []
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=False),
            start=header_row + 1,
        ):
            ym_value = self._extract_year_month(row, year_month_col)
            if ym_value is None:
                continue

            row_data: dict[str, float | None] = {}
            for col_idx, field_name in col_map.items():
                if col_idx < len(row):
                    row_data[field_name] = _to_float(row[col_idx].value)

            if all(v is None for v in row_data.values()):
                continue

            try:
                fund_count_raw = row_data.get("fund_count")
                fund_count = int(fund_count_raw) if fund_count_raw is not None else None
                record = ProductClassRecord(
                    product_class=sheet_name,
                    year_month=ym_value,
                    net_assets=row_data.get("net_assets"),
                    fund_count=fund_count,
                )
                records.append(record)
            except Exception as exc:
                logger.warning(
                    "Failed to parse B-2 row",
                    sheet=sheet_name,
                    row=row_idx,
                    error=str(exc),
                )
        return records

    # ------------------------------------------------------------------
    # B-3 (management company) parsing
    # ------------------------------------------------------------------

    def _find_header_row_b3(
        self,
        sheet: Worksheet,
    ) -> tuple[int, dict[int, str]]:
        """Find the header row for B-3 management company data.

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet to search.

        Returns
        -------
        tuple[int, dict[int, str]]
            Header row number (1-based) and column mapping.

        Raises
        ------
        ParseError
            If no header row is found.
        """
        header_keywords: dict[str, str] = {
            "会社名": "company_name",
            "会社": "company_name",
            "純資産": "net_assets",
            "本数": "fund_count",
            "ファンド": "fund_count",
        }

        for row_idx, row in enumerate(
            sheet.iter_rows(max_row=30, values_only=False), start=1
        ):
            matched_fields = self._match_header_cells(row, header_keywords)
            if "company_name" in matched_fields.values() and len(matched_fields) >= 2:
                logger.debug(
                    "B-3 header row found",
                    row=row_idx,
                    matched_columns=len(matched_fields),
                )
                return row_idx, matched_fields

        msg = "B-3 header row not found in worksheet"
        raise ParseError(
            msg,
            source=sheet.title or "unknown",
            reason="No matching header row for B-3 data",
        )

    def _extract_year_month_from_header_area(
        self,
        sheet: Worksheet,
        header_row: int,
    ) -> str | None:
        """Extract year-month from the header area rows above the data.

        Scans title rows above the header for patterns like
        "2024年1月末" or "2024/1".

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet.
        header_row : int
            1-based header row number.

        Returns
        -------
        str | None
            Year-month in "YYYY-MM" format, or None.
        """
        ym_title_re = re.compile(r"(\d{4})[年/](\d{1,2})")
        for row in sheet.iter_rows(min_row=1, max_row=header_row, values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                m = ym_title_re.search(text)
                if m:
                    result = _parse_ym_parts(int(m.group(1)), int(m.group(2)))
                    if result is not None:
                        return result
        return None

    def parse_b3(self, path: Path) -> list[ManagementCompanyRecord]:
        """Parse B-3 (management company) Excel file.

        Reads the first sheet and extracts company-level records.
        Year-month is determined from either a dedicated column
        or the header area text (e.g. "2024年1月末").

        Parameters
        ----------
        path : Path
            Path to the B-3 XLSX file.

        Returns
        -------
        list[ManagementCompanyRecord]
            Parsed management company records.

        Raises
        ------
        ParseError
            If the file cannot be read or parsed.
        """
        logger.info("Parsing B-3 management company Excel", path=str(path))
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise ParseError(
                f"Failed to open workbook: {exc}",
                source=str(path),
                reason=str(exc),
            ) from exc

        try:
            sheet = wb.worksheets[0] if wb.worksheets else None
            if sheet is None:
                raise ParseError(
                    "No active sheet found",
                    source=str(path),
                    reason="Workbook has no active sheet",
                )

            header_row, col_map = self._find_header_row_b3(sheet)

            # Check for year-month column in data rows
            year_month_col = self._detect_year_month_column(sheet, header_row, col_map)

            # If no year-month column, extract from header area
            fallback_ym: str | None = None
            if year_month_col is None:
                fallback_ym = self._extract_year_month_from_header_area(
                    sheet, header_row
                )

            records = self._read_b3_rows(
                sheet, header_row, col_map, year_month_col, fallback_ym
            )
        finally:
            wb.close()

        logger.info(
            "B-3 parsing complete",
            record_count=len(records),
            path=str(path),
        )
        return records

    def _read_b3_rows(
        self,
        sheet: Worksheet,
        header_row: int,
        col_map: dict[int, str],
        year_month_col: int | None,
        fallback_ym: str | None,
    ) -> list[ManagementCompanyRecord]:
        """Read data rows from B-3 sheet and convert to records.

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet.
        header_row : int
            1-based header row number.
        col_map : dict[int, str]
            Column index -> field name mapping.
        year_month_col : int | None
            Detected year-month column index.
        fallback_ym : str | None
            Fallback year-month from header area.

        Returns
        -------
        list[ManagementCompanyRecord]
            Parsed records.
        """
        records: list[ManagementCompanyRecord] = []
        company_col: int | None = None
        for col_idx, field_name in col_map.items():
            if field_name == "company_name":
                company_col = col_idx
                break

        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=False),
            start=header_row + 1,
        ):
            # Extract company name
            company_name: str | None = None
            if company_col is not None and company_col < len(row):
                val = row[company_col].value
                if val is not None:
                    company_name = str(val).strip()
            if not company_name:
                continue

            # Extract year-month
            ym_value: str | None = None
            if year_month_col is not None:
                ym_value = self._extract_year_month(row, year_month_col)
            if ym_value is None:
                ym_value = fallback_ym
            if ym_value is None:
                continue

            # Extract numeric fields
            row_data: dict[str, float | None] = {}
            for col_idx, field_name in col_map.items():
                if field_name != "company_name" and col_idx < len(row):
                    row_data[field_name] = _to_float(row[col_idx].value)

            try:
                fund_count_raw = row_data.get("fund_count")
                fund_count = int(fund_count_raw) if fund_count_raw is not None else None
                record = ManagementCompanyRecord(
                    company_name=company_name,
                    year_month=ym_value,
                    net_assets=row_data.get("net_assets"),
                    fund_count=fund_count,
                )
                records.append(record)
            except Exception as exc:
                logger.warning(
                    "Failed to parse B-3 row",
                    row=row_idx,
                    error=str(exc),
                )
        return records

    # ------------------------------------------------------------------
    # A-2 (overall status) parsing
    # ------------------------------------------------------------------

    def _find_header_row_a2(
        self,
        sheet: Worksheet,
    ) -> tuple[int, dict[int, str]]:
        """Find the header row for A-2 overall status data.

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet to search.

        Returns
        -------
        tuple[int, dict[int, str]]
            Header row number (1-based) and column mapping.

        Raises
        ------
        ParseError
            If no header row is found.
        """
        header_keywords: dict[str, str] = {
            "純資産": "total_net_assets",
            "本数": "total_fund_count",
            "ファンド": "total_fund_count",
            "設定": "total_inflow",
            "解約": "total_outflow",
        }

        for row_idx, row in enumerate(
            sheet.iter_rows(max_row=30, values_only=False), start=1
        ):
            matched_fields = self._match_header_cells(row, header_keywords)
            if len(matched_fields) >= 2:
                logger.debug(
                    "A-2 header row found",
                    row=row_idx,
                    matched_columns=len(matched_fields),
                )
                return row_idx, matched_fields

        msg = "A-2 header row not found in worksheet"
        raise ParseError(
            msg,
            source=sheet.title or "unknown",
            reason="No matching header row for A-2 data",
        )

    def parse_a2(self, path: Path) -> list[OverallStatusRecord]:
        """Parse A-2 (overall status) Excel file.

        Reads the first sheet of the workbook and converts each
        data row into an ``OverallStatusRecord``.

        Parameters
        ----------
        path : Path
            Path to the A-2 XLSX file.

        Returns
        -------
        list[OverallStatusRecord]
            Parsed overall status records.

        Raises
        ------
        ParseError
            If the file cannot be read or parsed.
        """
        logger.info("Parsing A-2 overall status Excel", path=str(path))
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise ParseError(
                f"Failed to open workbook: {exc}",
                source=str(path),
                reason=str(exc),
            ) from exc

        try:
            sheet = wb.worksheets[0] if wb.worksheets else None
            if sheet is None:
                raise ParseError(
                    "No active sheet found",
                    source=str(path),
                    reason="Workbook has no active sheet",
                )

            header_row, col_map = self._find_header_row_a2(sheet)
            year_month_col = self._detect_year_month_column(sheet, header_row, col_map)
            records = self._read_a2_rows(sheet, header_row, col_map, year_month_col)
        finally:
            wb.close()

        logger.info(
            "A-2 parsing complete",
            record_count=len(records),
            path=str(path),
        )
        return records

    def _read_a2_rows(
        self,
        sheet: Worksheet,
        header_row: int,
        col_map: dict[int, str],
        year_month_col: int | None,
    ) -> list[OverallStatusRecord]:
        """Read data rows from A-2 sheet and convert to records.

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet.
        header_row : int
            1-based header row number.
        col_map : dict[int, str]
            Column index -> field name mapping.
        year_month_col : int | None
            Detected year-month column index.

        Returns
        -------
        list[OverallStatusRecord]
            Parsed records.
        """
        records: list[OverallStatusRecord] = []
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=False),
            start=header_row + 1,
        ):
            ym_value = self._extract_year_month(row, year_month_col)
            if ym_value is None:
                continue

            row_data: dict[str, float | None] = {}
            for col_idx, field_name in col_map.items():
                if col_idx < len(row):
                    row_data[field_name] = _to_float(row[col_idx].value)

            if all(v is None for v in row_data.values()):
                continue

            try:
                fund_count_raw = row_data.get("total_fund_count")
                fund_count = int(fund_count_raw) if fund_count_raw is not None else None
                record = OverallStatusRecord(
                    year_month=ym_value,
                    total_net_assets=row_data.get("total_net_assets"),
                    total_fund_count=fund_count,
                    total_inflow=row_data.get("total_inflow"),
                    total_outflow=row_data.get("total_outflow"),
                )
                records.append(record)
            except Exception as exc:
                logger.warning(
                    "Failed to parse A-2 row",
                    row=row_idx,
                    error=str(exc),
                )
        return records
