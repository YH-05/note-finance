"""Parser for NISA growth investment target Excel files.

Reads XLSX files using openpyxl, maps columns via the dictionaries
in ``fund_db.config.constants``, and returns Pydantic model instances.

Classes
-------
NisaParser
    Parses NISA target fund Excel files.

Examples
--------
>>> from pathlib import Path
>>> parser = NisaParser()
>>> # funds = parser.parse_unlisted(Path(".tmp/unlisted_fund_for_investor.xlsx"))
"""

from __future__ import annotations

from typing import TYPE_CHECKING, Any

import openpyxl

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

from fund_db._logging import get_logger
from fund_db.config.constants import NISA_LISTED_COLUMNS, NISA_UNLISTED_COLUMNS
from fund_db.exceptions import ParseError
from fund_db.nisa.models import NisaListedEtf, NisaUnlistedFund

logger = get_logger(__name__, module="nisa_parser")


def _normalize_value(value: Any) -> str | None:
    """Normalize a cell value to a string or None.

    Parameters
    ----------
    value : Any
        Raw cell value from openpyxl.

    Returns
    -------
    str | None
        Stripped string if non-empty, otherwise None.
    """
    if value is None:
        return None
    text = str(value).strip()
    if text in {"", "None"}:
        return None
    return text


class NisaParser:
    """Parses NISA target fund Excel files.

    Reads XLSX workbooks, locates header rows by matching column names
    from ``NISA_UNLISTED_COLUMNS`` / ``NISA_LISTED_COLUMNS``, and
    converts each data row into a Pydantic model instance.

    Examples
    --------
    >>> parser = NisaParser()
    """

    def _find_header_row(
        self,
        sheet: Worksheet,
        column_mapping: dict[str, str],
    ) -> tuple[int, dict[int, str]]:
        """Find the header row and build column index mapping.

        Parameters
        ----------
        sheet : Worksheet
            openpyxl worksheet to search.
        column_mapping : dict[str, str]
            Mapping from Japanese column names to field names.

        Returns
        -------
        tuple[int, dict[int, str]]
            Header row number and mapping from column index to field name.

        Raises
        ------
        ParseError
            If no header row is found.
        """
        jp_names = set(column_mapping.keys())

        for row_idx, row in enumerate(
            sheet.iter_rows(max_row=30, values_only=False), start=1
        ):
            cell_values = {
                str(cell.value).strip() for cell in row if cell.value is not None
            }
            matched = jp_names & cell_values
            if len(matched) >= len(jp_names) // 2 + 1:
                # Build column index -> field name mapping
                col_map: dict[int, str] = {}
                for cell in row:
                    if cell.value is not None:
                        header_text = str(cell.value).strip()
                        if header_text in column_mapping:
                            col_map[cell.column - 1] = column_mapping[header_text]
                logger.debug(
                    "Header row found",
                    row=row_idx,
                    matched_columns=len(col_map),
                )
                return row_idx, col_map

        msg = "Header row not found in worksheet"
        raise ParseError(
            msg, source=sheet.title or "unknown", reason="No matching header row"
        )

    def parse_unlisted(self, path: Path) -> list[NisaUnlistedFund]:
        """Parse non-listed investment trust Excel file.

        Parameters
        ----------
        path : Path
            Path to the XLSX file.

        Returns
        -------
        list[NisaUnlistedFund]
            Parsed fund records.

        Raises
        ------
        ParseError
            If the file cannot be read or parsed.
        """
        logger.info("Parsing unlisted fund Excel", path=str(path))
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise ParseError(
                f"Failed to open workbook: {exc}",
                source=str(path),
                reason=str(exc),
            ) from exc

        try:
            sheet = wb.active
            if sheet is None:
                raise ParseError(
                    "No active sheet found",
                    source=str(path),
                    reason="Workbook has no active sheet",
                )

            header_row, col_map = self._find_header_row(sheet, NISA_UNLISTED_COLUMNS)

            records: list[NisaUnlistedFund] = []
            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=False),
                start=header_row + 1,
            ):
                row_data: dict[str, str | None] = {}
                for col_idx, field_name in col_map.items():
                    cell = row[col_idx]
                    row_data[field_name] = _normalize_value(cell.value)

                # Skip entirely empty rows
                if all(v is None for v in row_data.values()):
                    continue

                # Required fields check
                code = row_data.get("association_code")
                name = row_data.get("fund_name")
                company = row_data.get("management_company")
                if not code or not name or not company:
                    logger.debug(
                        "Skipping row with missing required fields",
                        row=row_idx,
                    )
                    continue

                try:
                    fund = NisaUnlistedFund(
                        association_code=code,
                        fund_name=name,
                        management_company=company,
                        asset_class=row_data.get("asset_class"),
                        investment_region=row_data.get("investment_region"),
                        fund_type=row_data.get("fund_type"),
                        benchmark_index=row_data.get("benchmark_index"),
                        expense_ratio=row_data.get("expense_ratio"),
                        tsumitate_eligible=row_data.get("tsumitate_eligible"),
                        growth_eligible=row_data.get("growth_eligible"),
                    )
                    records.append(fund)
                except Exception as exc:
                    logger.warning(
                        "Failed to parse row",
                        row=row_idx,
                        error=str(exc),
                    )
        finally:
            wb.close()

        logger.info(
            "Unlisted fund parsing complete",
            record_count=len(records),
            path=str(path),
        )
        return records

    def parse_listed(self, path: Path) -> list[NisaListedEtf]:
        """Parse listed ETF/REIT Excel file.

        Parameters
        ----------
        path : Path
            Path to the XLSX file.

        Returns
        -------
        list[NisaListedEtf]
            Parsed ETF/REIT records.

        Raises
        ------
        ParseError
            If the file cannot be read or parsed.
        """
        logger.info("Parsing listed ETF Excel", path=str(path))
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise ParseError(
                f"Failed to open workbook: {exc}",
                source=str(path),
                reason=str(exc),
            ) from exc

        try:
            sheet = wb.active
            if sheet is None:
                raise ParseError(
                    "No active sheet found",
                    source=str(path),
                    reason="Workbook has no active sheet",
                )

            header_row, col_map = self._find_header_row(sheet, NISA_LISTED_COLUMNS)

            records: list[NisaListedEtf] = []
            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=False),
                start=header_row + 1,
            ):
                row_data: dict[str, str | None] = {}
                for col_idx, field_name in col_map.items():
                    cell = row[col_idx]
                    row_data[field_name] = _normalize_value(cell.value)

                # Skip entirely empty rows
                if all(v is None for v in row_data.values()):
                    continue

                # Required fields check
                code = row_data.get("ticker_code")
                name = row_data.get("fund_name")
                if not code or not name:
                    logger.debug(
                        "Skipping row with missing required fields",
                        row=row_idx,
                    )
                    continue

                try:
                    etf = NisaListedEtf(
                        ticker_code=code,
                        fund_name=name,
                        management_company=row_data.get("management_company"),
                        benchmark_index=row_data.get("benchmark_index"),
                        expense_ratio=row_data.get("expense_ratio"),
                        trading_unit=row_data.get("trading_unit"),
                    )
                    records.append(etf)
                except Exception as exc:
                    logger.warning(
                        "Failed to parse row",
                        row=row_idx,
                        error=str(exc),
                    )
        finally:
            wb.close()

        logger.info(
            "Listed ETF parsing complete",
            record_count=len(records),
            path=str(path),
        )
        return records
