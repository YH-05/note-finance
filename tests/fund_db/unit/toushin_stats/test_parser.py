"""Unit tests for fund_db.toushin_stats.parser module.

Tests parse_b1(), parse_b2(), parse_b3(), parse_a2() using synthetic
openpyxl workbooks.
"""

from __future__ import annotations

from datetime import datetime
from typing import TYPE_CHECKING

import openpyxl
import pytest

if TYPE_CHECKING:
    from pathlib import Path

from fund_db.exceptions import ParseError
from fund_db.toushin_stats.parser import ToushinStatsParser, _to_float, _to_year_month

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def parser() -> ToushinStatsParser:
    """Create a ToushinStatsParser instance."""
    return ToushinStatsParser()


@pytest.fixture
def b1_workbook(tmp_path: Path) -> Path:
    """Create a synthetic B-1 Excel workbook for testing.

    Mimics the structure of toushin_B1_shisan_zougen.xlsx:
    - Title rows at top
    - Header row with column names
    - Data rows with year_month and numeric values
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "資産増減状況"

    # Title rows
    ws.cell(row=1, column=1, value="投資信託の資産増減状況")
    ws.cell(row=2, column=1, value="（単位：百万円）")

    # Header row
    ws.cell(row=3, column=1, value="年月")
    ws.cell(row=3, column=2, value="純資産総額")
    ws.cell(row=3, column=3, value="設定額")
    ws.cell(row=3, column=4, value="解約額")
    ws.cell(row=3, column=5, value="純増減額")

    # Data rows
    ws.cell(row=4, column=1, value="2024/1")
    ws.cell(row=4, column=2, value=1500000.0)
    ws.cell(row=4, column=3, value=50000.0)
    ws.cell(row=4, column=4, value=30000.0)
    ws.cell(row=4, column=5, value=20000.0)

    ws.cell(row=5, column=1, value="2024/2")
    ws.cell(row=5, column=2, value=1520000.0)
    ws.cell(row=5, column=3, value=55000.0)
    ws.cell(row=5, column=4, value=35000.0)
    ws.cell(row=5, column=5, value=20000.0)

    ws.cell(row=6, column=1, value="2024/3")
    ws.cell(row=6, column=2, value=1550000.0)
    ws.cell(row=6, column=3, value=60000.0)
    ws.cell(row=6, column=4, value=32000.0)
    ws.cell(row=6, column=5, value=28000.0)

    path = tmp_path / "test_b1.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def b1_workbook_with_none_cells(tmp_path: Path) -> Path:
    """Create a B-1 workbook with None/empty cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "資産増減状況"

    # Header row
    ws.cell(row=1, column=1, value="年月")
    ws.cell(row=1, column=2, value="純資産総額")
    ws.cell(row=1, column=3, value="設定額")
    ws.cell(row=1, column=4, value="解約額")
    ws.cell(row=1, column=5, value="純増減額")

    # Data row with some None values
    ws.cell(row=2, column=1, value="2024/1")
    ws.cell(row=2, column=2, value=1500000.0)
    ws.cell(row=2, column=3, value=None)
    ws.cell(row=2, column=4, value=None)
    ws.cell(row=2, column=5, value=20000.0)

    # Empty row (should be skipped)
    ws.cell(row=3, column=1, value=None)
    ws.cell(row=3, column=2, value=None)

    # Another valid row
    ws.cell(row=4, column=1, value="2024/2")
    ws.cell(row=4, column=2, value=1520000.0)
    ws.cell(row=4, column=3, value=55000.0)
    ws.cell(row=4, column=4, value=35000.0)
    ws.cell(row=4, column=5, value=20000.0)

    path = tmp_path / "test_b1_none.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def b1_workbook_datetime_format(tmp_path: Path) -> Path:
    """Create a B-1 workbook with datetime year_month values."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    ws.cell(row=1, column=1, value="年月")
    ws.cell(row=1, column=2, value="純資産総額")
    ws.cell(row=1, column=3, value="設定額")

    # Use datetime objects for year_month
    ws.cell(row=2, column=1, value=datetime(2024, 1, 1))
    ws.cell(row=2, column=2, value=1500000.0)
    ws.cell(row=2, column=3, value=50000.0)

    ws.cell(row=3, column=1, value=datetime(2024, 6, 15))
    ws.cell(row=3, column=2, value=1600000.0)
    ws.cell(row=3, column=3, value=60000.0)

    path = tmp_path / "test_b1_datetime.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def empty_workbook(tmp_path: Path) -> Path:
    """Create an empty workbook with no data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Empty"

    path = tmp_path / "test_empty.xlsx"
    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Tests: _to_float helper
# ---------------------------------------------------------------------------


class TestToFloat:
    """Tests for the _to_float helper function."""

    def test_正常系_整数を変換できる(self) -> None:
        assert _to_float(42) == 42.0

    def test_正常系_浮動小数点を変換できる(self) -> None:
        assert _to_float(3.14) == 3.14

    def test_正常系_文字列数値を変換できる(self) -> None:
        assert _to_float("1500000") == 1500000.0

    def test_正常系_カンマ区切り文字列を変換できる(self) -> None:
        assert _to_float("1,500,000") == 1500000.0

    def test_正常系_NoneをNoneに変換する(self) -> None:
        assert _to_float(None) is None

    def test_正常系_空文字列をNoneに変換する(self) -> None:
        assert _to_float("") is None

    def test_正常系_ハイフンをNoneに変換する(self) -> None:
        assert _to_float("-") is None

    def test_正常系_NAをNoneに変換する(self) -> None:
        assert _to_float("N/A") is None

    def test_正常系_NaN浮動小数点をNoneに変換する(self) -> None:
        assert _to_float(float("nan")) is None

    def test_正常系_Inf浮動小数点をNoneに変換する(self) -> None:
        assert _to_float(float("inf")) is None

    def test_正常系_非数値文字列をNoneに変換する(self) -> None:
        assert _to_float("abc") is None


# ---------------------------------------------------------------------------
# Tests: _to_year_month helper
# ---------------------------------------------------------------------------


class TestToYearMonth:
    """Tests for the _to_year_month helper function."""

    def test_正常系_YYYY_MM形式を変換できる(self) -> None:
        assert _to_year_month("2024-01") == "2024-01"

    def test_正常系_YYYYスラッシュM形式を変換できる(self) -> None:
        assert _to_year_month("2024/1") == "2024-01"

    def test_正常系_YYYYスラッシュMM形式を変換できる(self) -> None:
        assert _to_year_month("2024/01") == "2024-01"

    def test_正常系_YYYY年M月形式を変換できる(self) -> None:
        assert _to_year_month("2024年1月") == "2024-01"

    def test_正常系_YYYY年MM月形式を変換できる(self) -> None:
        assert _to_year_month("2024年12月") == "2024-12"

    def test_正常系_datetimeオブジェクトを変換できる(self) -> None:
        dt = datetime(2024, 6, 15, 10, 30)
        assert _to_year_month(dt) == "2024-06"

    def test_正常系_NoneをNoneに変換する(self) -> None:
        assert _to_year_month(None) is None

    def test_正常系_空文字列をNoneに変換する(self) -> None:
        assert _to_year_month("") is None

    def test_正常系_不正な文字列をNoneに変換する(self) -> None:
        assert _to_year_month("abc") is None

    def test_正常系_ハイフンをNoneに変換する(self) -> None:
        assert _to_year_month("-") is None


# ---------------------------------------------------------------------------
# Tests: ToushinStatsParser.parse_b1
# ---------------------------------------------------------------------------


class TestParseB1:
    """Tests for ToushinStatsParser.parse_b1()."""

    def test_正常系_合成ワークブックからレコードを読み込める(
        self,
        parser: ToushinStatsParser,
        b1_workbook: Path,
    ) -> None:
        records = parser.parse_b1(b1_workbook)
        assert len(records) == 3

    def test_正常系_最初のレコードのフィールド値が正しい(
        self,
        parser: ToushinStatsParser,
        b1_workbook: Path,
    ) -> None:
        records = parser.parse_b1(b1_workbook)
        first = records[0]
        assert first.year_month == "2024-01"
        assert first.net_assets == 1500000.0
        assert first.inflow == 50000.0
        assert first.outflow == 30000.0
        assert first.net_flow == 20000.0

    def test_正常系_3レコード目のフィールド値が正しい(
        self,
        parser: ToushinStatsParser,
        b1_workbook: Path,
    ) -> None:
        records = parser.parse_b1(b1_workbook)
        third = records[2]
        assert third.year_month == "2024-03"
        assert third.net_assets == 1550000.0
        assert third.inflow == 60000.0
        assert third.outflow == 32000.0
        assert third.net_flow == 28000.0

    def test_正常系_Noneセルを含むレコードを正しく変換できる(
        self,
        parser: ToushinStatsParser,
        b1_workbook_with_none_cells: Path,
    ) -> None:
        records = parser.parse_b1(b1_workbook_with_none_cells)
        # Should have 2 records (empty row skipped)
        assert len(records) == 2
        first = records[0]
        assert first.year_month == "2024-01"
        assert first.net_assets == 1500000.0
        assert first.inflow is None
        assert first.outflow is None
        assert first.net_flow == 20000.0

    def test_正常系_datetime形式のyear_monthを正しく変換できる(
        self,
        parser: ToushinStatsParser,
        b1_workbook_datetime_format: Path,
    ) -> None:
        records = parser.parse_b1(b1_workbook_datetime_format)
        assert len(records) == 2
        assert records[0].year_month == "2024-01"
        assert records[1].year_month == "2024-06"

    def test_異常系_存在しないファイルでParseError(
        self,
        parser: ToushinStatsParser,
        tmp_path: Path,
    ) -> None:
        with pytest.raises(ParseError):
            parser.parse_b1(tmp_path / "nonexistent.xlsx")

    def test_異常系_ヘッダー行なしでParseError(
        self,
        parser: ToushinStatsParser,
        empty_workbook: Path,
    ) -> None:
        with pytest.raises(ParseError):
            parser.parse_b1(empty_workbook)

    def test_正常系_空データ行のみの場合は空リスト(
        self,
        parser: ToushinStatsParser,
        tmp_path: Path,
    ) -> None:
        """Workbook with header but no data rows returns empty list."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.cell(row=1, column=1, value="年月")
        ws.cell(row=1, column=2, value="純資産総額")
        ws.cell(row=1, column=3, value="設定額")
        path = tmp_path / "test_header_only.xlsx"
        wb.save(path)
        wb.close()

        records = parser.parse_b1(path)
        assert records == []

    def test_正常系_レコードがAssetFlowRecord型である(
        self,
        parser: ToushinStatsParser,
        b1_workbook: Path,
    ) -> None:
        from fund_db.toushin_stats.models import AssetFlowRecord

        records = parser.parse_b1(b1_workbook)
        for record in records:
            assert isinstance(record, AssetFlowRecord)

    def test_正常系_レコードをmodel_dumpで辞書に変換できる(
        self,
        parser: ToushinStatsParser,
        b1_workbook: Path,
    ) -> None:
        records = parser.parse_b1(b1_workbook)
        data = records[0].model_dump()
        assert isinstance(data, dict)
        assert "year_month" in data
        assert "net_assets" in data


# ---------------------------------------------------------------------------
# Fixtures: B-2, B-3, A-2
# ---------------------------------------------------------------------------


@pytest.fixture
def b2_workbook(tmp_path: Path) -> Path:
    """Create a synthetic B-2 Excel workbook with multiple sheets.

    Mimics the structure of toushin_B2_shohin_bunrui.xlsx:
    - Multiple sheets, each representing a product class
    - Each sheet has year_month + numeric columns (net_assets, fund_count)
    """
    wb = openpyxl.Workbook()

    # Remove default sheet
    default_ws = wb.active
    assert default_ws is not None

    # Sheet 1: 株式投信
    ws1 = wb.create_sheet(title="株式投信")
    ws1.cell(row=1, column=1, value="商品分類別純資産総額等")
    ws1.cell(row=2, column=1, value="（単位：百万円）")
    ws1.cell(row=3, column=1, value="年月")
    ws1.cell(row=3, column=2, value="純資産総額")
    ws1.cell(row=3, column=3, value="ファンド本数")
    ws1.cell(row=4, column=1, value="2024/1")
    ws1.cell(row=4, column=2, value=800000.0)
    ws1.cell(row=4, column=3, value=150)
    ws1.cell(row=5, column=1, value="2024/2")
    ws1.cell(row=5, column=2, value=820000.0)
    ws1.cell(row=5, column=3, value=155)

    # Sheet 2: 公社債投信
    ws2 = wb.create_sheet(title="公社債投信")
    ws2.cell(row=1, column=1, value="商品分類別純資産総額等")
    ws2.cell(row=2, column=1, value="（単位：百万円）")
    ws2.cell(row=3, column=1, value="年月")
    ws2.cell(row=3, column=2, value="純資産総額")
    ws2.cell(row=3, column=3, value="ファンド本数")
    ws2.cell(row=4, column=1, value="2024/1")
    ws2.cell(row=4, column=2, value=300000.0)
    ws2.cell(row=4, column=3, value=80)

    # Sheet 3: MMF
    ws3 = wb.create_sheet(title="MMF")
    ws3.cell(row=1, column=1, value="商品分類別純資産総額等")
    ws3.cell(row=2, column=1, value="（単位：百万円）")
    ws3.cell(row=3, column=1, value="年月")
    ws3.cell(row=3, column=2, value="純資産総額")
    ws3.cell(row=3, column=3, value="ファンド本数")
    ws3.cell(row=4, column=1, value="2024/1")
    ws3.cell(row=4, column=2, value=50000.0)
    ws3.cell(row=4, column=3, value=10)

    # Remove the default sheet
    wb.remove(default_ws)

    path = tmp_path / "test_b2.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def b3_workbook(tmp_path: Path) -> Path:
    """Create a synthetic B-3 Excel workbook for testing.

    Mimics the structure of toushin_B3_unyou_kaisha.xlsx:
    - Single sheet with management company rows
    - Columns: company_name, net_assets, fund_count
    - Year-month from a header area or column
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "運用会社別"

    # Title rows
    ws.cell(row=1, column=1, value="運用会社別純資産総額等")
    ws.cell(row=2, column=1, value="2024年1月末")
    ws.cell(row=3, column=1, value="（単位：百万円）")

    # Header row
    ws.cell(row=4, column=1, value="会社名")
    ws.cell(row=4, column=2, value="純資産総額")
    ws.cell(row=4, column=3, value="ファンド本数")

    # Data rows - management companies
    ws.cell(row=5, column=1, value="野村アセットマネジメント")
    ws.cell(row=5, column=2, value=500000.0)
    ws.cell(row=5, column=3, value=200)

    ws.cell(row=6, column=1, value="大和アセットマネジメント")
    ws.cell(row=6, column=2, value=400000.0)
    ws.cell(row=6, column=3, value=180)

    ws.cell(row=7, column=1, value="三菱UFJアセットマネジメント")
    ws.cell(row=7, column=2, value=350000.0)
    ws.cell(row=7, column=3, value=160)

    path = tmp_path / "test_b3.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def b3_workbook_with_ym_column(tmp_path: Path) -> Path:
    """Create a B-3 workbook with an explicit year_month column."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "運用会社別"

    # Header row
    ws.cell(row=1, column=1, value="年月")
    ws.cell(row=1, column=2, value="会社名")
    ws.cell(row=1, column=3, value="純資産総額")
    ws.cell(row=1, column=4, value="ファンド本数")

    # Data rows
    ws.cell(row=2, column=1, value="2024/1")
    ws.cell(row=2, column=2, value="野村アセットマネジメント")
    ws.cell(row=2, column=3, value=500000.0)
    ws.cell(row=2, column=4, value=200)

    ws.cell(row=3, column=1, value="2024/1")
    ws.cell(row=3, column=2, value="大和アセットマネジメント")
    ws.cell(row=3, column=3, value=400000.0)
    ws.cell(row=3, column=4, value=180)

    path = tmp_path / "test_b3_ym_col.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def a2_workbook(tmp_path: Path) -> Path:
    """Create a synthetic A-2 Excel workbook for testing.

    Mimics the structure of toushin_A2_zentaizo.xlsx:
    - Single sheet with monthly time series
    - Columns: year_month, total_net_assets, total_fund_count,
               total_inflow, total_outflow
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "全体像"

    # Title rows
    ws.cell(row=1, column=1, value="投資信託の全体像")
    ws.cell(row=2, column=1, value="（単位：百万円）")

    # Header row
    ws.cell(row=3, column=1, value="年月")
    ws.cell(row=3, column=2, value="純資産総額")
    ws.cell(row=3, column=3, value="ファンド本数")
    ws.cell(row=3, column=4, value="設定額")
    ws.cell(row=3, column=5, value="解約額")

    # Data rows
    ws.cell(row=4, column=1, value="2024/1")
    ws.cell(row=4, column=2, value=2000000.0)
    ws.cell(row=4, column=3, value=500)
    ws.cell(row=4, column=4, value=100000.0)
    ws.cell(row=4, column=5, value=80000.0)

    ws.cell(row=5, column=1, value="2024/2")
    ws.cell(row=5, column=2, value=2050000.0)
    ws.cell(row=5, column=3, value=510)
    ws.cell(row=5, column=4, value=110000.0)
    ws.cell(row=5, column=5, value=85000.0)

    ws.cell(row=6, column=1, value="2024/3")
    ws.cell(row=6, column=2, value=2100000.0)
    ws.cell(row=6, column=3, value=520)
    ws.cell(row=6, column=4, value=120000.0)
    ws.cell(row=6, column=5, value=90000.0)

    path = tmp_path / "test_a2.xlsx"
    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Tests: ToushinStatsParser.parse_b2
# ---------------------------------------------------------------------------


class TestParseB2:
    """Tests for ToushinStatsParser.parse_b2()."""

    def test_正常系_複数シートからレコードを読み込める(
        self,
        parser: ToushinStatsParser,
        b2_workbook: Path,
    ) -> None:
        records = parser.parse_b2(b2_workbook)
        assert len(records) == 4  # 2 (株式投信) + 1 (公社債投信) + 1 (MMF)

    def test_正常系_シート名がproduct_classに設定される(
        self,
        parser: ToushinStatsParser,
        b2_workbook: Path,
    ) -> None:
        records = parser.parse_b2(b2_workbook)
        product_classes = {r.product_class for r in records}
        assert "株式投信" in product_classes
        assert "公社債投信" in product_classes
        assert "MMF" in product_classes

    def test_正常系_最初のシートのレコード値が正しい(
        self,
        parser: ToushinStatsParser,
        b2_workbook: Path,
    ) -> None:
        records = parser.parse_b2(b2_workbook)
        kabushiki_records = [r for r in records if r.product_class == "株式投信"]
        assert len(kabushiki_records) == 2
        first = kabushiki_records[0]
        assert first.year_month == "2024-01"
        assert first.net_assets == 800000.0
        assert first.fund_count == 150

    def test_正常系_2番目のシートのレコード値が正しい(
        self,
        parser: ToushinStatsParser,
        b2_workbook: Path,
    ) -> None:
        records = parser.parse_b2(b2_workbook)
        koshasai_records = [r for r in records if r.product_class == "公社債投信"]
        assert len(koshasai_records) == 1
        first = koshasai_records[0]
        assert first.year_month == "2024-01"
        assert first.net_assets == 300000.0
        assert first.fund_count == 80

    def test_正常系_レコードがProductClassRecord型である(
        self,
        parser: ToushinStatsParser,
        b2_workbook: Path,
    ) -> None:
        from fund_db.toushin_stats.models import ProductClassRecord

        records = parser.parse_b2(b2_workbook)
        for record in records:
            assert isinstance(record, ProductClassRecord)

    def test_異常系_存在しないファイルでParseError(
        self,
        parser: ToushinStatsParser,
        tmp_path: Path,
    ) -> None:
        with pytest.raises(ParseError):
            parser.parse_b2(tmp_path / "nonexistent.xlsx")

    def test_正常系_ヘッダーなしシートはスキップされる(
        self,
        parser: ToushinStatsParser,
        tmp_path: Path,
    ) -> None:
        """A sheet with no recognizable header should be skipped."""
        wb = openpyxl.Workbook()
        default_ws = wb.active
        assert default_ws is not None

        # Sheet with valid data
        ws1 = wb.create_sheet(title="株式投信")
        ws1.cell(row=1, column=1, value="年月")
        ws1.cell(row=1, column=2, value="純資産総額")
        ws1.cell(row=1, column=3, value="ファンド本数")
        ws1.cell(row=2, column=1, value="2024/1")
        ws1.cell(row=2, column=2, value=100000.0)
        ws1.cell(row=2, column=3, value=50)

        # Sheet with no recognizable header
        ws2 = wb.create_sheet(title="メモ")
        ws2.cell(row=1, column=1, value="このシートはデータなし")

        wb.remove(default_ws)
        path = tmp_path / "test_b2_skip.xlsx"
        wb.save(path)
        wb.close()

        records = parser.parse_b2(path)
        assert len(records) == 1
        assert records[0].product_class == "株式投信"


# ---------------------------------------------------------------------------
# Tests: ToushinStatsParser.parse_b3
# ---------------------------------------------------------------------------


class TestParseB3:
    """Tests for ToushinStatsParser.parse_b3()."""

    def test_正常系_運用会社レコードを読み込める(
        self,
        parser: ToushinStatsParser,
        b3_workbook: Path,
    ) -> None:
        records = parser.parse_b3(b3_workbook)
        assert len(records) == 3

    def test_正常系_最初の運用会社レコードの値が正しい(
        self,
        parser: ToushinStatsParser,
        b3_workbook: Path,
    ) -> None:
        records = parser.parse_b3(b3_workbook)
        first = records[0]
        assert first.company_name == "野村アセットマネジメント"
        assert first.net_assets == 500000.0
        assert first.fund_count == 200

    def test_正常系_year_monthがヘッダー領域から推定される(
        self,
        parser: ToushinStatsParser,
        b3_workbook: Path,
    ) -> None:
        records = parser.parse_b3(b3_workbook)
        # year_month should be extracted from the title area "2024年1月末"
        for record in records:
            assert record.year_month == "2024-01"

    def test_正常系_year_monthカラムが存在する場合(
        self,
        parser: ToushinStatsParser,
        b3_workbook_with_ym_column: Path,
    ) -> None:
        records = parser.parse_b3(b3_workbook_with_ym_column)
        assert len(records) == 2
        assert records[0].year_month == "2024-01"
        assert records[0].company_name == "野村アセットマネジメント"

    def test_正常系_レコードがManagementCompanyRecord型である(
        self,
        parser: ToushinStatsParser,
        b3_workbook: Path,
    ) -> None:
        from fund_db.toushin_stats.models import ManagementCompanyRecord

        records = parser.parse_b3(b3_workbook)
        for record in records:
            assert isinstance(record, ManagementCompanyRecord)

    def test_異常系_存在しないファイルでParseError(
        self,
        parser: ToushinStatsParser,
        tmp_path: Path,
    ) -> None:
        with pytest.raises(ParseError):
            parser.parse_b3(tmp_path / "nonexistent.xlsx")

    def test_異常系_ヘッダー行なしでParseError(
        self,
        parser: ToushinStatsParser,
        empty_workbook: Path,
    ) -> None:
        with pytest.raises(ParseError):
            parser.parse_b3(empty_workbook)


# ---------------------------------------------------------------------------
# Tests: ToushinStatsParser.parse_a2
# ---------------------------------------------------------------------------


class TestParseA2:
    """Tests for ToushinStatsParser.parse_a2()."""

    def test_正常系_全体統計レコードを読み込める(
        self,
        parser: ToushinStatsParser,
        a2_workbook: Path,
    ) -> None:
        records = parser.parse_a2(a2_workbook)
        assert len(records) == 3

    def test_正常系_最初のレコードの値が正しい(
        self,
        parser: ToushinStatsParser,
        a2_workbook: Path,
    ) -> None:
        records = parser.parse_a2(a2_workbook)
        first = records[0]
        assert first.year_month == "2024-01"
        assert first.total_net_assets == 2000000.0
        assert first.total_fund_count == 500
        assert first.total_inflow == 100000.0
        assert first.total_outflow == 80000.0

    def test_正常系_3レコード目の値が正しい(
        self,
        parser: ToushinStatsParser,
        a2_workbook: Path,
    ) -> None:
        records = parser.parse_a2(a2_workbook)
        third = records[2]
        assert third.year_month == "2024-03"
        assert third.total_net_assets == 2100000.0
        assert third.total_fund_count == 520
        assert third.total_inflow == 120000.0
        assert third.total_outflow == 90000.0

    def test_正常系_レコードがOverallStatusRecord型である(
        self,
        parser: ToushinStatsParser,
        a2_workbook: Path,
    ) -> None:
        from fund_db.toushin_stats.models import OverallStatusRecord

        records = parser.parse_a2(a2_workbook)
        for record in records:
            assert isinstance(record, OverallStatusRecord)

    def test_異常系_存在しないファイルでParseError(
        self,
        parser: ToushinStatsParser,
        tmp_path: Path,
    ) -> None:
        with pytest.raises(ParseError):
            parser.parse_a2(tmp_path / "nonexistent.xlsx")

    def test_異常系_ヘッダー行なしでParseError(
        self,
        parser: ToushinStatsParser,
        empty_workbook: Path,
    ) -> None:
        with pytest.raises(ParseError):
            parser.parse_a2(empty_workbook)

    def test_正常系_Noneセルを含むレコードを正しく変換できる(
        self,
        parser: ToushinStatsParser,
        tmp_path: Path,
    ) -> None:
        """A-2 with some None values in numeric columns."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None

        ws.cell(row=1, column=1, value="年月")
        ws.cell(row=1, column=2, value="純資産総額")
        ws.cell(row=1, column=3, value="ファンド本数")
        ws.cell(row=1, column=4, value="設定額")
        ws.cell(row=1, column=5, value="解約額")

        ws.cell(row=2, column=1, value="2024/1")
        ws.cell(row=2, column=2, value=2000000.0)
        ws.cell(row=2, column=3, value=None)
        ws.cell(row=2, column=4, value=100000.0)
        ws.cell(row=2, column=5, value=None)

        path = tmp_path / "test_a2_none.xlsx"
        wb.save(path)
        wb.close()

        records = parser.parse_a2(path)
        assert len(records) == 1
        assert records[0].total_net_assets == 2000000.0
        assert records[0].total_fund_count is None
        assert records[0].total_inflow == 100000.0
        assert records[0].total_outflow is None
