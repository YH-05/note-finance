"""Unit tests for fund_db.nisa.parser module.

Tests Excel parsing with real .tmp/ files (skipped if files are absent)
and synthetic in-memory workbooks for deterministic validation.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from fund_db._utils import normalize_cell_value
from fund_db.exceptions import ParseError
from fund_db.nisa.models import NisaListedEtf, NisaUnlistedFund
from fund_db.nisa.parser import NisaParser

# ---------------------------------------------------------------------------
# Paths to real Excel files (skipped if absent)
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[4]
TMP_UNLISTED = PROJECT_ROOT / ".tmp" / "unlisted_fund_for_investor.xlsx"
TMP_LISTED = PROJECT_ROOT / ".tmp" / "listed_fund_for_investor.xlsx"


# ---------------------------------------------------------------------------
# Helper: create synthetic workbooks in tmp_path
# ---------------------------------------------------------------------------


def _create_unlisted_workbook(path: Path) -> Path:
    """Create a minimal XLSX with unlisted fund data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    # Header row
    headers = [
        "協会コード",
        "ファンド名称",
        "運用会社名",
        "投資対象資産",
        "投資対象地域",
        "インデックス型/アクティブ型",
        "対象インデックス",
        "信託報酬（税込）",
        "つみたて投資枠",
        "成長投資枠",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Data rows
    data_rows = [
        [
            "01311046",
            "eMAXIS Slim 全世界株式",
            "三菱UFJアセット",
            "株式",
            "内外",
            "インデックス型",
            "MSCI ACWI",
            "0.05775%以内",
            "○",
            "○",
        ],
        [
            "01312001",
            "iFree 日経225",
            "大和アセット",
            "株式",
            "国内",
            "インデックス型",
            "日経225",
            "0.154%",
            "○",
            "",
        ],
        [
            "01313099",
            "アクティブファンドA",
            "テスト運用",
            "株式",
            "海外",
            "アクティブ型",
            None,
            "1.10%",
            "",
            "○",
        ],
    ]
    for row_idx, data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    xlsx_path = path / "test_unlisted.xlsx"
    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


def _create_listed_workbook(path: Path) -> Path:
    """Create a minimal XLSX with listed ETF data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    headers = [
        "銘柄コード",
        "銘柄名称",
        "管理会社",
        "対象インデックス",
        "信託報酬（税込）",
        "売買単位",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    data_rows = [
        ["1306", "TOPIX連動型上場投資信託", "野村AM", "TOPIX", "0.0968%以内", "10口"],
        ["1321", "日経225連動型上場投資信託", "野村AM", "日経225", "0.198%以内", "1口"],
    ]
    for row_idx, data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    xlsx_path = path / "test_listed.xlsx"
    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


def _create_empty_workbook(path: Path) -> Path:
    """Create an empty XLSX file."""
    wb = openpyxl.Workbook()
    xlsx_path = path / "empty.xlsx"
    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


# ---------------------------------------------------------------------------
# Tests: _normalize_value
# ---------------------------------------------------------------------------


class TestNormalizeValue:
    """Tests for the _normalize_value helper function."""

    def test_正常系_文字列をそのまま返す(self) -> None:
        assert normalize_cell_value("hello") == "hello"

    def test_正常系_数値を文字列に変換する(self) -> None:
        assert normalize_cell_value(42) == "42"

    def test_正常系_Noneを返す(self) -> None:
        assert normalize_cell_value(None) is None

    def test_正常系_空文字列をNoneに変換する(self) -> None:
        assert normalize_cell_value("") is None

    def test_正常系_空白文字列をNoneに変換する(self) -> None:
        assert normalize_cell_value("   ") is None

    def test_正常系_前後空白をトリムする(self) -> None:
        assert normalize_cell_value("  hello  ") == "hello"

    def test_正常系_None文字列をNoneに変換する(self) -> None:
        assert normalize_cell_value("None") is None

    def test_正常系_浮動小数点を文字列に変換する(self) -> None:
        assert normalize_cell_value(0.05775) == "0.05775"


# ---------------------------------------------------------------------------
# Tests: NisaParser with synthetic workbooks
# ---------------------------------------------------------------------------


class TestNisaParserSynthetic:
    """Tests for NisaParser using synthetic workbooks."""

    def test_正常系_非上場ファンドをパースできる(self, tmp_path: Path) -> None:
        xlsx_path = _create_unlisted_workbook(tmp_path)
        parser = NisaParser()
        funds = parser.parse_unlisted(xlsx_path)
        assert len(funds) == 3
        assert all(isinstance(f, NisaUnlistedFund) for f in funds)
        assert funds[0].association_code == "01311046"
        assert funds[0].fund_name == "eMAXIS Slim 全世界株式"
        assert funds[0].management_company == "三菱UFJアセット"
        assert funds[0].asset_class == "株式"
        assert funds[0].tsumitate_eligible == "○"

    def test_正常系_上場ETFをパースできる(self, tmp_path: Path) -> None:
        xlsx_path = _create_listed_workbook(tmp_path)
        parser = NisaParser()
        etfs = parser.parse_listed(xlsx_path)
        assert len(etfs) == 2
        assert all(isinstance(e, NisaListedEtf) for e in etfs)
        assert etfs[0].ticker_code == "1306"
        assert etfs[0].fund_name == "TOPIX連動型上場投資信託"
        assert etfs[0].management_company == "野村AM"
        assert etfs[0].benchmark_index == "TOPIX"
        assert etfs[0].trading_unit == "10口"

    def test_正常系_空文字列がNoneに変換される(self, tmp_path: Path) -> None:
        xlsx_path = _create_unlisted_workbook(tmp_path)
        parser = NisaParser()
        funds = parser.parse_unlisted(xlsx_path)
        # Row 2 has empty tsumitate_eligible ""
        fund_ifree = funds[1]
        assert fund_ifree.tsumitate_eligible == "○"
        assert fund_ifree.growth_eligible is None

    def test_正常系_Noneのbenchmark_indexが保持される(self, tmp_path: Path) -> None:
        xlsx_path = _create_unlisted_workbook(tmp_path)
        parser = NisaParser()
        funds = parser.parse_unlisted(xlsx_path)
        # Row 3 has benchmark_index=None
        active_fund = funds[2]
        assert active_fund.benchmark_index is None

    def test_異常系_ヘッダー行が見つからないとParseError(self, tmp_path: Path) -> None:
        xlsx_path = _create_empty_workbook(tmp_path)
        parser = NisaParser()
        with pytest.raises(ParseError, match="Header row not found"):
            parser.parse_unlisted(xlsx_path)

    def test_異常系_存在しないファイルでParseError(self) -> None:
        parser = NisaParser()
        with pytest.raises(ParseError, match="Failed to open workbook"):
            parser.parse_unlisted(Path("/nonexistent/file.xlsx"))

    def test_異常系_上場ETFでヘッダー行なしParseError(self, tmp_path: Path) -> None:
        xlsx_path = _create_empty_workbook(tmp_path)
        parser = NisaParser()
        with pytest.raises(ParseError, match="Header row not found"):
            parser.parse_listed(xlsx_path)


# ---------------------------------------------------------------------------
# Tests: NisaParser with real Excel files (skipped if absent)
# ---------------------------------------------------------------------------


@pytest.mark.skipif(
    not TMP_UNLISTED.exists(), reason="テスト用 Excel ファイルが存在しません"
)
class TestNisaParserRealUnlisted:
    """Tests with real unlisted fund Excel from .tmp/."""

    def test_正常系_実ファイルから2200件以上パースできる(self) -> None:
        parser = NisaParser()
        funds = parser.parse_unlisted(TMP_UNLISTED)
        assert len(funds) >= 2200
        assert all(isinstance(f, NisaUnlistedFund) for f in funds)

    def test_正常系_全レコードに必須フィールドが存在する(self) -> None:
        parser = NisaParser()
        funds = parser.parse_unlisted(TMP_UNLISTED)
        for fund in funds:
            assert fund.association_code
            assert fund.fund_name
            assert fund.management_company


@pytest.mark.skipif(
    not TMP_LISTED.exists(), reason="テスト用 Excel ファイルが存在しません"
)
class TestNisaParserRealListed:
    """Tests with real listed ETF Excel from .tmp/."""

    def test_正常系_実ファイルから390件以上パースできる(self) -> None:
        parser = NisaParser()
        etfs = parser.parse_listed(TMP_LISTED)
        assert len(etfs) >= 390
        assert all(isinstance(e, NisaListedEtf) for e in etfs)

    def test_正常系_全レコードに必須フィールドが存在する(self) -> None:
        parser = NisaParser()
        etfs = parser.parse_listed(TMP_LISTED)
        for etf in etfs:
            assert etf.ticker_code
            assert etf.fund_name
